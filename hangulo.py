import sys
import re
import zipfile
import os
import shutil
import time
import configparser
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QProgressDialog, QMessageBox
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject
from openpyxl import load_workbook

# Initialize an empty list to store the strings
invalid_file_list = []

hangulo_folder = ".hangulo"

def get_config_value(key):
    """
    주어진 key에 해당하는 값을 config 파일에서 가져오는 함수입니다.

    Parameters:
        key (str): 찾을 값의 key

    Returns:
        str: key에 해당하는 값 (찾을 수 없으면 None 반환)
    """
    config = configparser.ConfigParser()
    try:
        # Open the 'config.ini' file with 'utf-8' encoding
        with open('config.ini', 'r', encoding='utf-8') as file:
            config.read_file(file)
    except UnicodeDecodeError as e:
        print(f"An error occurred while reading 'config.ini': {e}")
        # Handle the error or use a different encoding if necessary


    value = None
    if 'Section1' in config and key in config['Section1']:
        value = config['Section1'][key]

    return value

def number_to_korean_amount(number):
    """
    주어진 숫자를 한글로 변환하여 금액 표현과 함께 리턴하는 함수입니다.

    Parameters:
        number (int): 변환할 숫자

    Returns:
        str: 한글로 변환된 금액 표현 (예: "128,600원(금일십이만팔천육백원정)")
    """
    korean_numbers = ["일", "이", "삼", "사", "오", "육", "칠", "팔", "구"]
    korean_units = ["", "십", "백", "천"]
    korean_big_units = ["", "만", "억", "조", "경", "해", "경", "조", "억", "만"]

    number_str = str(number)
    result = []
    length = len(number_str)

    for i, digit in enumerate(number_str):
        num = int(digit)
        unit_index = (length - 1 - i) % 4
        if num != 0:
            if i > 0 and unit_index == 0 and number_str[i] == '1':
                result.append(korean_numbers[0])
            else:
                result.append(korean_numbers[num - 1])
            result.append(korean_units[unit_index])
        if unit_index == 0 and i < length - 1:
            result.append(korean_big_units[(length - 1 - i) // 4])

    return "금" + "".join(result) + "원정"

def replace_values_in_xml(xlsx_file, xml_file):
    # Read XML file
    with open(xml_file, 'rt', encoding='UTF-8') as file:
        xml_content = file.read()

    pattern = r'%(\w+)%'
    # print(csv_file)
    # print(xml_content.encode('utf-8').decode('utf-8'))
    # sys.exit(0)

    allMatches = re.findall(pattern, xml_content)
    # Convert the list to a set to remove duplicates, and then back to a list
    matches = list(set(allMatches))

    if len(matches) == 0:
        raise ValueError(f"템플릿 Hwp 파일 '{xml_file}'에서 '%변수%'가 존재하지 않습니다.")
    variables = []

    # 엑셀 파일 불러오기
    wb = load_workbook(xlsx_file, data_only=True)

    # 기본으로 첫 번째 시트 선택
    ws = wb.active

    total_amount = 0
    for key in matches:
        value = get_config_value(key)
        variables.append(value)
        if value is not None:
            count = xml_content.count(f"%{key}%")

            pattern_to_replace = f"%{key}%"

            for i in range(2, count + 2):
                cell_name = f"{value}{i}"
                cell_value = ws[cell_name].value
                if (key == '법정기일'):
                    cell_value_str = cell_value.date().strftime('%Y-%m-%d')
                elif (key == '세액' or key == '가산금' or key == '계'):
                    cell_value = ws[cell_name].value
                    if str(cell_value).isdigit():
                        # 숫자로 이루어진 문자열이라면 정수형으로 변환
                        amount = int(cell_value)
                        cell_value_str = '{:,}'.format(amount)
                    else:
                        raise ValueError(f"XLSX 파일에서 해당 셀의 값이 숫자가 아닙니다. config.ini 파일에서 '{key}' 항목의 셀 이름을 확인하세요.")

                    if (key == '계'):
                        total_amount += amount
                    
                else:
                    cell_value_str = str(cell_value)
                print(f'{cell_name}: {cell_value_str}')
                xml_content = xml_content.replace(pattern_to_replace, cell_value_str, 1)

    # Calculate the sum of amounts from L2 to L(num_tax_numbers + 1)
    tax_total_amount = "{:,}원".format(total_amount)
    korean_total_amount_str = number_to_korean_amount(total_amount)
    tax_total_amount_str = tax_total_amount + '(' + korean_total_amount_str  + ')'

    xml_content = xml_content.replace("%TAX_TOTAL_AMOUNT_STR%", tax_total_amount_str)
    xml_content = xml_content.replace("%TAX_TOTAL_AMOUNT%", tax_total_amount)
    # xml_content = re.sub(re.escape("%TAX_TOTAL_AMOUNT_STR%"), korean_amount_num)
    # xml_content = re.sub(re.escape("%TAX_TOTAL_AMOUNT%"), korean_amount_num, xml_content)

    return xml_content

class MessageSignal(QObject):
    message_signal = pyqtSignal(str, str, str)

class ConverterThread(QThread):
    progress_signal = pyqtSignal(int)

    def __init__(self, directory):
        super().__init__()
        self.directory = directory
        self.message_signal = MessageSignal()

    def run(self):
        # Check if the .hangulo folder exists and is empty
        if os.path.exists(hangulo_folder):
            try:
                # Remove the .hangulo folder
                shutil.rmtree(hangulo_folder)
                print(f"The '{hangulo_folder}' folder has been deleted.")
            except Exception as e:
                print(f"Error occurred while deleting the folder: {e}")
        else:
            print(f"The '{hangulo_folder}' folder does not exist.")


        xlsx_files = [filename for filename in os.listdir(self.directory) if filename.endswith('.xlsx')]
        if not xlsx_files:
            self.message_signal.message_signal.emit('No XLSX Files', 'There are no .xlsx files in the selected directory.', 'warning')
            return

        total_files = len(xlsx_files)
        for i, filename in enumerate(xlsx_files):
            file_path = os.path.join(self.directory, filename)

            print("### ", file_path)
        
            # 엑셀 파일 불러오기
            wb = load_workbook(file_path)

            # 기본으로 첫 번째 시트 선택
            ws = wb.active

            # 마지막 행 번호 가져오기
            num_tax_numbers = ws.max_row - 1

            if (num_tax_numbers > 5):
                invalid_file_list.append(filename)
                continue

            hwpx_file = f'template-tax-{num_tax_numbers}.hwpx'
            hwpx_file_name = os.path.splitext(hwpx_file)[0]
            zip_file = f'{hwpx_file_name}.zip'
            
            extract_to = '.hangulo'
            if not os.path.exists(extract_to):
                os.makedirs(extract_to)
            
            zip_file_path = os.path.join(extract_to, zip_file)
            shutil.copy(hwpx_file, zip_file_path)

            with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
                hwpx_dir = os.path.join(extract_to, hwpx_file_name)
                if not os.path.exists(hwpx_dir):
                    zip_ref.extractall(hwpx_dir)

            contents_dir = os.path.join(hwpx_dir, 'Contents')
            xml_file = 'section0.xml'

            xml_file_path = os.path.join(contents_dir , xml_file)

            try:
                xml_output_result = replace_values_in_xml(file_path, xml_file_path)
            except ValueError as e:
                self.message_signal.message_signal.emit('변환 실패', str(e), 'warning')
                sys.exit(1)

            # 디렉토리를 복사합니다.
            gen_hwpx_file_name = os.path.splitext(filename)[0]
            gen_hwpx_dir = os.path.join(self.directory, gen_hwpx_file_name)

            # 디렉토리가 존재하는지 확인합니다.
            if os.path.exists(gen_hwpx_dir):
                # 디렉토리를 삭제합니다.
                try:
                    shutil.rmtree(gen_hwpx_dir)
                except OSError as e:
                    self.message_signal.message_signal.emit('변환 실패', f"기존에 생성된 디렉토리 '{gen_hwpx_dir}'를 삭제할 수 없습니다. '{gen_hwpx_dir}' 를 삭제한 뒤 다시 시도해 주세요. 오류: {e}", 'warning')
                    sys.exit(1)
            shutil.copytree(hwpx_dir, gen_hwpx_dir)
            gen_contents_dir = os.path.join(gen_hwpx_dir, 'Contents')
            gen_xml_file_path = os.path.join(gen_contents_dir , xml_file)

            with open(gen_xml_file_path, 'wt', encoding='UTF-8') as output_file:
                output_file.write(xml_output_result)
    
            # 디렉토리를 zip 파일로 압축합니다.
            gen_zip_file_name_path = os.path.join(self.directory, gen_hwpx_file_name)

            gen_zip_file_path = f'{gen_zip_file_name_path}.zip'

            gen_hwpx_file = f'{gen_hwpx_file_name}.hwpx'
            gen_hwpx_file_path = os.path.join(self.directory, gen_hwpx_file)

            # zip 파일이 존재하는지 확인합니다.
            if os.path.exists(gen_zip_file_path):
                # 파일을 삭제합니다.
                os.remove(gen_zip_file_path)

            # Hwpx 파일이 존재하는지 확인합니다.
            if os.path.exists(gen_hwpx_file_path):
                # 파일 이름을 삭제합니다.
                try:
                    os.remove(gen_hwpx_file_path)
                except OSError as e:
                    self.message_signal.message_signal.emit('변환 실패', f"기존에 생성된 Hwp 파일 '{gen_hwpx_file_path}'를 삭제할 수 없습니다. 열린 '{gen_hwpx_file_path}' 를 닫은 뒤 다시 시도해 주세요. 오류: {e}", 'warning')
                    sys.exit(1)
                

            shutil.make_archive(gen_zip_file_name_path, 'zip', gen_hwpx_dir)

            # 디렉토리가 존재하는지 확인합니다.
            if os.path.exists(gen_hwpx_dir):
                # 디렉토리를 삭제합니다.
                try:
                    shutil.rmtree(gen_hwpx_dir)
                except OSError as e:
                    self.message_signal.message_signal.emit('변환 실패', f"기존에 생성된 디렉토리 '{gen_hwpx_dir}'를 삭제할 수 없습니다. '{gen_hwpx_dir}' 를 삭제한 뒤 다시 시도해 주세요. 오류: {e}", 'warning')
                    sys.exit(1)

            shutil.copy(gen_zip_file_path, gen_hwpx_file_path)

            # zip 파일이 존재하는지 확인합니다.
            if os.path.exists(gen_zip_file_path):
                # 파일을 삭제합니다.
                os.remove(gen_zip_file_path)

            time.sleep(1)  # 컨버팅 시뮬레이션을 위한 딜레이
            progress_percent = int((i + 1) / total_files * 100)
            self.progress_signal.emit(progress_percent)

        appended_invalid_files = ''
        for i, invalid_file in enumerate(invalid_file_list):
            if i > 0:
                appended_invalid_files += ', '
            appended_invalid_files += invalid_file

        if (appended_invalid_files != ''):
            self.message_signal.message_signal.emit('Hwp로 변환 완료', f"선택된 폴더 내의 xlsx 파일들을 모두 변환 하였지만 (변환된 Hwp 폴더 위치: '{self.directory}'), '{appended_invalid_files}' 파일들은 6개 이상 건수를 가지고 있어 변환할 수 없습니다!", 'warning')
        else:
            self.message_signal.message_signal.emit('Hwp로 변환 완료', f"변환 완료하였습니다. (변환 Hwp 위치: '{self.directory}')", 'info')

class ConverterApp(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.converter_thread = None

        self.message_signal = MessageSignal()
        self.message_signal.message_signal.connect(self.show_message)

    def show_message(self, title, message, message_type):
        if message_type == 'warning':
            QMessageBox.warning(self, title, message)
        else:
            QMessageBox.information(self, title, message)

    def cancel_conversion(self):
        if self.converter_thread:
            self.converter_thread.requestInterruption()


    def select_directory(self):
        options = QFileDialog.Options()
        directory = QFileDialog.getExistingDirectory(self, 'Select Directory')
        if not directory:
             QMessageBox.information(None, 'Application Closed', 'You have canceled the selection. The application will now exit.')
             sys.exit(0)
        if directory:
            print(f'Selected directory: {directory}')
            self.progress_dialog = QProgressDialog('Converting...', 'Cancel', 0, 100, self)
            self.progress_dialog.setWindowTitle("Hwp로 변환");
            self.progress_dialog.setMinimumWidth(300);
            self.progress_dialog.setWindowModality(Qt.WindowModal)
            self.progress_dialog.canceled.connect(self.cancel_conversion)
            # Disable the Cancel button
            self.progress_dialog.setCancelButton(None)
            self.progress_dialog.setValue(0)
            self.progress_dialog.show()
            self.converter_thread = ConverterThread(directory)
            self.converter_thread.progress_signal.connect(self.update_progress)
            self.converter_thread.message_signal.message_signal.connect(self.show_message)
            self.converter_thread.finished.connect(self.conversion_completed)

            self.converter_thread.start()

    def update_progress(self, value):
        self.progress_dialog.setValue(value)

    def conversion_completed(self):
        self.progress_dialog.hide()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    converter_app = ConverterApp()
    converter_app.select_directory()  # Moved select_directory() out of __init__()
    sys.exit(app.exec_())

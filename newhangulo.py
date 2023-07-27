# 엑셀 파일 불러오기
workbook = load_workbook('파일명.xlsx')

# 기본으로 첫 번째 시트 선택
sheet = workbook.active

# 마지막 행 번호 가져오기
last_row = sheet.max_row


# 마지막 행 데이터 가져오기
last_row_data = sheet[f"A{last_row}":"Z{last_row}"]

# 마지막 행 데이터 출력 (각 셀 값 출력)
for row in last_row_data:
    for cell in row:
        print(cell.value)







##################






from openpyxl import load_workbook

wb = load_workbook('simple Data.xlsx')
data = wb.active

print(data['A1'].value)
print(data['A2'].value)
print(data['B1'].value)
print(data['B2'].value)






#########################




